import os
import sys
import json
from typing import Any, Dict, List

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

import openpyxl  # type: ignore
from sqlalchemy.orm import Session
from app.database import SessionLocal
from app.models import UploadedDocument, Application, FormData

EXCEL_PATH = os.path.join(PROJECT_ROOT, "data", "BoardCertificate_License_Schema.xlsx")
SHEET_NAME = "Medical Certificate"

OCR_COLUMNS = [
    "Issuer", "Recipient Name", "Title/Degree", "Field of Study",
    "Date of Certification", "Signatories", "Seal Detected",
    "Document Type", "Certificate No."
]
VERIFICATION_COLUMNS = [
    "Demo_Verification_Attribute1 (PDF Format Match)", "Comment 1",
    "Demo_Verification_Attribute2 (Board Certificate Match)", "Comment 2"
]
TRACE_COLUMNS = [
    "Flag", "Row_ID", "provider_first_name", "provider_last_name",
    "provider_name", "npi"
]

def _s(val: Any) -> Any:
    if val is None:
        return None
    try:
        import datetime as _dt
        if isinstance(val, (_dt.date, _dt.datetime)):
            return val.isoformat()
    except Exception:
        pass
    if isinstance(val, (str, int, float, bool)):
        return val
    return str(val)

def detect_header_row(ws) -> int:
    max_scan = min(8, ws.max_row)
    for r in range(1, max_scan + 1):
        vals = [(str(c.value) or "").strip().lower() for c in ws[r]]
        if any("npi" in v for v in vals) or any("issuer" in v for v in vals):
            return r
    return 1

def load_rows(path: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in Excel file {path}")
    ws = wb[SHEET_NAME]
    header_row = detect_header_row(ws)
    headers_by_idx: Dict[int, str] = {}
    for j, cell in enumerate(ws[header_row], start=1):
        headers_by_idx[j] = (str(cell.value) or "").strip()
    rows: List[Dict[str, Any]] = []
    for i in range(header_row + 1, ws.max_row + 1):
        rec: Dict[str, Any] = {}
        empty = True
        for j, key in headers_by_idx.items():
            v = ws.cell(row=i, column=j).value
            rec[key] = v
            if v not in (None, ""):
                empty = False
        if not empty:
            rows.append(rec)
    return rows

def build_payload(rec: Dict[str, Any]) -> Dict[str, Any]:
    ocr = {col: _s(rec.get(col)) for col in OCR_COLUMNS if rec.get(col) not in (None, "")}
    verification = {col: _s(rec.get(col)) for col in VERIFICATION_COLUMNS if rec.get(col) not in (None, "")}
    npi_val = rec.get("npi")
    npi_str = None
    if npi_val is not None:
        try:
            npi_str = str(int(npi_val)) if isinstance(npi_val, (int, float)) else str(npi_val)
            npi_str = npi_str.strip()
        except Exception:
            npi_str = str(npi_val).strip()
    return {
        "source": f"{os.path.basename(EXCEL_PATH)}:{SHEET_NAME}",
        "npi": npi_str,
        "ocr": ocr,
        "verification": verification,
    }

def upsert_medical_certificate(db: Session, form_id: str, payload: Dict[str, Any]):
    row = (
        db.query(UploadedDocument)
        .filter(UploadedDocument.form_id == form_id)
        .filter(UploadedDocument.file_type == "MEDICAL_TRAINING_CERTIFICATE")
        .first()
    )
    if not row:
        row = UploadedDocument(
            form_id=form_id,
            filename=f"medical_cert_{form_id}.json",
            file_extension="json",
            file_type="MEDICAL_TRAINING_CERTIFICATE",
            status="In Progress",
        )
        db.add(row)
        db.flush()
    row.ocr_output = json.dumps(payload.get("ocr", {}))
    row.verification_data = json.dumps(payload.get("verification", {}))
    if not row.status:
        row.status = "In Progress"
    return row

def main():
    rows = load_rows(EXCEL_PATH)
    print(f"Loaded Medical Certificate rows: {len(rows)}")

    db: Session = SessionLocal()
    try:
        npi_to_form: Dict[str, str] = {}
        for f in db.query(FormData).all():
            if f.npi and f.form_id:
                npi_to_form[str(f.npi).strip()] = f.form_id
        for a in db.query(Application).all():
            if a.npi and a.form_id:
                npi_to_form.setdefault(str(a.npi).strip(), a.form_id)

        upserts = 0
        for rec in rows:
            npi_val = rec.get("npi")
            if not npi_val:
                continue
            try:
                npi = str(int(npi_val)) if isinstance(npi_val, (int, float)) else str(npi_val)
                npi = npi.strip()
            except Exception:
                npi = str(npi_val).strip()
            form_id = npi_to_form.get(npi)
            if not form_id:
                continue
            payload = build_payload(rec)
            upsert_medical_certificate(db, form_id, payload)
            upserts += 1
        db.commit()
        print(f"Medical Certificate ingest complete. Upserts={upserts}")
    finally:
        db.close()

if __name__ == "__main__":
    main()
