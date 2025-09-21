import os
import sys
import json
from typing import Dict, Any

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

import openpyxl
from sqlalchemy.orm import Session
from app.database import SessionLocal
from app.models import UploadedDocument, Application, FormData

SHEET_NAME = "Malpractice Insurance"

OCR_HEADERS = [
    "Insured Name",
    "Insurer Name",
    "Policy Number",
    "Policy Effective Date",
    "Policy Expiration Date",
    "Liability Limit (Per Claim)",
    "Liability Limit (Aggregate)",
]
VERIF_HEADERS = [
    "Demo_Verification_Attribute1 (PDF Format Match)",
    "Comment 1",
    "Demo_Verification_Attribute2 (DEA Verification)",
]


def detect_header_row(ws) -> int:
    for r in range(1, min(7, ws.max_row + 1)):
        row_vals = [(str(c.value) or "").strip().lower() for c in ws[r]]
        if any("insured name" in v for v in row_vals):
            return r
    return 2


def load_rows(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in Excel file")
    ws = wb[SHEET_NAME]
    header_row = detect_header_row(ws)
    headers_by_idx = {}
    for j, cell in enumerate(ws[header_row], start=1):
        headers_by_idx[j] = (str(cell.value) or "").strip()
    rows = []
    for i in range(header_row + 1, ws.max_row + 1):
        rec = {}
        for j, key in headers_by_idx.items():
            rec[key] = ws.cell(row=i, column=j).value
        rows.append(rec)
    return rows


def _s(val: Any) -> Any:
    # Convert date/time and other non-JSON-serializable types to string
    # Keep None as None
    if val is None:
        return None
    # Normalize floats/ints representing dates or numbers into strings as-is
    try:
        import datetime as _dt
        if isinstance(val, (_dt.date, _dt.datetime)):
            return val.isoformat()
    except Exception:
        pass
    # Leave simple JSON-safe types unchanged
    if isinstance(val, (str, int, float, bool)):
        return val
    # Fallback to string
    return str(val)


def structured_payload(rec: Dict[str, Any]) -> Dict[str, Any]:
    ocr = {h: _s(rec.get(h)) for h in OCR_HEADERS}
    ver = {
        "pdf_format_match": _s(rec.get("Demo_Verification_Attribute1 (PDF Format Match)")) or _s(rec.get("PDF Format Match")),
        "comment_1": _s(rec.get("Comment 1")),
        "verification_status": _s(rec.get("Demo_Verification_Attribute2 (DEA Verification)")) or _s(rec.get("Verification Status")),
    }
    return {
        "source": "Other_Attributes_Schema.xlsx:Malpractice Insurance",
        "ocr": ocr,
        "verification": ver,
    }


def upsert_malpractice(db: Session, form_id: str, payload: Dict[str, Any]):
    row = (
        db.query(UploadedDocument)
        .filter(UploadedDocument.form_id == form_id)
        .filter(UploadedDocument.file_type == "malpractice_insurance")
        .first()
    )
    if not row:
        row = UploadedDocument(
            form_id=form_id,
            filename=f"malpractice_{form_id}.json",
            file_extension="json",
            file_type="malpractice_insurance",
            status="In Progress",
        )
        db.add(row)
        db.flush()
    row.ocr_output = json.dumps(payload.get("ocr", {}))
    row.verification_data = json.dumps(payload.get("verification", {}))
    if not row.status:
        row.status = "In Progress"
    return row


def main():
    excel_path = os.path.join(PROJECT_ROOT, "data", "Other_Attributes_Schema.xlsx")
    rows = load_rows(excel_path)
    print(f"Loaded Malpractice rows: {len(rows)}")

    db: Session = SessionLocal()
    try:
        # Build indices
        npi_to_form: Dict[str, str] = {}
        for f in db.query(FormData).all():
            if f.npi and f.form_id:
                npi_to_form[str(f.npi).strip()] = f.form_id
        for a in db.query(Application).all():
            if a.npi and a.form_id:
                npi_to_form.setdefault(str(a.npi).strip(), a.form_id)

        name_to_form: Dict[str, str] = {}
        for a in db.query(Application).all():
            if a.form_id and a.name:
                name_to_form.setdefault(a.name.strip().lower(), a.form_id)
        for f in db.query(FormData).all():
            if f.form_id and f.provider_name:
                name_to_form.setdefault(f.provider_name.strip().lower(), f.form_id)

        upserts = 0
        for rec in rows:
            # Prefer NPI linkage when available
            npi_val = None
            for k, v in rec.items():
                if "npi" in (k or "").strip().lower():
                    npi_val = v
                    break
            form_id = None
            if npi_val is not None:
                try:
                    npi_str = str(int(npi_val)) if isinstance(npi_val, (int, float)) else str(npi_val)
                    npi_str = npi_str.strip()
                except Exception:
                    npi_str = str(npi_val).strip()
                form_id = npi_to_form.get(npi_str)
            if not form_id:
                insured_name = (str(rec.get("Insured Name") or "").strip().lower())
                form_id = name_to_form.get(insured_name)
            if not form_id:
                continue
            payload = structured_payload(rec)
            upsert_malpractice(db, form_id, payload)
            upserts += 1
        db.commit()
        print(f"Malpractice Insurance ingest complete. Upserts={upserts}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
