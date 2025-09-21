import os, sys, json
from typing import Any, Dict

# Ensure project root path
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

import openpyxl
from sqlalchemy.orm import Session
from app.database import SessionLocal
from app.models import UploadedDocument, Application, FormData

SHEET_NAME = "Sanctioned"

# Map spreadsheet headers to a normalized sanctions payload
HEADER_MAP = {
    # Expected columns (case-insensitive):
    "flag": "flag",
    "row_id": "row_id",
    "provider_first_name": "first_name",
    "provider_last_name": "last_name",
    "provider_name": "name",
    "provider_npi": "npi",
    "demo_verification_attribute1 (sanction status)": "status_attr",
    "comment 1": "comment_1",
    "demo_verification_attribute2 (sanction details)": "details_attr",
    "comment 2": "comment_2",
}


def load_sheet(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in Excel file")
    ws = wb[SHEET_NAME]
    # Build header index map (lowercased)
    headers = {}
    for j, cell in enumerate(ws[2], start=1):  # assume header row at 2
        key = (str(cell.value) or "").strip().lower()
        if key:
            headers[j] = key
    # Parse rows starting at 3
    rows = []
    for i in range(3, ws.max_row + 1):
        rec = {}
        for j, key in headers.items():
            rec[key] = ws.cell(row=i, column=j).value
        # basic filter: keep only flagged 'Sanctioned'
        flag = (str(rec.get("flag") or "").strip().lower())
        raw_npi = rec.get("provider_npi") or rec.get("npi")
        if isinstance(raw_npi, float) or isinstance(raw_npi, int):
            npi = str(int(raw_npi))
        else:
            npi = (str(raw_npi or "").strip())
        if not npi:
            continue
        if flag and flag.startswith("sanctioned"):
            rows.append(rec)
    return rows


def normalized_payload(rec: Dict[str, Any]) -> Dict[str, Any]:
    # Build a structured dict to store in UploadedDocument.verification_data
    raw_npi = rec.get("provider_npi") or rec.get("npi")
    npi_str = str(int(raw_npi)) if isinstance(raw_npi, (int, float)) else str(raw_npi or "").strip()
    return {
        "source": "Other_Attributes_Schema.xlsx:Sanctioned",
        "npi": npi_str,
        "provider": {
            "first_name": rec.get("provider_first_name") or rec.get("first_name"),
            "last_name": rec.get("provider_last_name") or rec.get("last_name"),
            "name": rec.get("provider_name") or rec.get("name"),
        },
        "sanction": {
            "status": rec.get("demo_verification_attribute1 (sanction status)") or rec.get("status_attr"),
            "details": rec.get("demo_verification_attribute2 (sanction details)") or rec.get("details_attr"),
            "comment_1": rec.get("comment 1") or rec.get("comment_1"),
            "comment_2": rec.get("comment 2") or rec.get("comment_2"),
        },
    }


def upsert_sanction(db: Session, form_id: str, payload: Dict[str, Any]):
    # Find an existing sanctions record for this form
    row = (
        db.query(UploadedDocument)
        .filter(UploadedDocument.form_id == form_id)
        .filter(UploadedDocument.file_type == "sanctions")
        .first()
    )
    if not row:
        row = UploadedDocument(
            form_id=form_id,
            filename=f"sanctions_{form_id}.json",
            file_extension="json",
            file_type="sanctions",
            status="In Progress",
        )
        db.add(row)
        db.flush()
    row.verification_data = json.dumps(payload)
    if not row.status:
        row.status = "In Progress"
    return row


def main():
    excel_path = os.path.join(PROJECT_ROOT, "data", "Other_Attributes_Schema.xlsx")
    records = load_sheet(excel_path)
    print(f"Loaded sanctioned records: {len(records)}")

    db: Session = SessionLocal()
    try:
        npi_to_form = {}
        # Build quick index: NPI -> form_id (prefer direct form_data match)
        for f in db.query(FormData).all():
            if f.npi:
                npi_to_form[str(f.npi).strip()] = f.form_id
        # Fallback via applications table
        for a in db.query(Application).all():
            if a.npi and a.form_id:
                npi_to_form.setdefault(str(a.npi).strip(), a.form_id)

        upserts = 0
        app_updates = 0
        for rec in records:
            raw_npi = rec.get("provider_npi") or rec.get("npi")
            npi = str(int(raw_npi)) if isinstance(raw_npi, (int, float)) else str(raw_npi or "").strip()
            form_id = npi_to_form.get(npi)
            if not form_id:
                continue
            payload = normalized_payload(rec)
            upsert_sanction(db, form_id, payload)
            upserts += 1
            # Mark application as SANCTIONED if NPI matches
            app = db.query(Application).filter(Application.form_id == form_id).first()
            if app:
                if app.psv_status != "SANCTIONED":
                    app.psv_status = "SANCTIONED"
                if app.committee_status != "SANCTIONED":
                    app.committee_status = "SANCTIONED"
                app_updates += 1
        db.commit()
        print(f"Ingest complete. Sanctions upserts={upserts} applications_marked={app_updates}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
