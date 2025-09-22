import os
import sys
import json
from typing import Any, Dict, List

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

import openpyxl  # type: ignore
from sqlalchemy.orm import Session
from app.database import SessionLocal
from app.models import UploadedDocument, Application, FormData

EXCEL_PATH = os.path.join(PROJECT_ROOT, "data", "BoardCertificate_License_Schema.xlsx")
SHEET_NAME = "Board_Certification"


def _s(val: Any) -> Any:
    if val is None:
        return None
    try:
        import datetime as _dt
        if isinstance(val, (_dt.date, _dt.datetime)):
            return val.isoformat()
    except Exception:
        pass
    if isinstance(val, (str, int, float, bool)):
        return val
    return str(val)


def detect_header_row(ws) -> int:
    # Look for a header that contains either NPI, provider_name, or Board_ABMS_*
    max_scan = min(8, ws.max_row)
    for r in range(1, max_scan + 1):
        vals = [(str(c.value) or "").strip().lower() for c in ws[r]]
        if any("npi" in v for v in vals) or any(v.startswith("board_abms_") for v in vals) or any("provider_name" in v for v in vals):
            return r
    return 1


def load_rows(path: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in Excel file {path}")
    ws = wb[SHEET_NAME]
    header_row = detect_header_row(ws)
    headers_by_idx: Dict[int, str] = {}
    for j, cell in enumerate(ws[header_row], start=1):
        headers_by_idx[j] = (str(cell.value) or "").strip()
    rows: List[Dict[str, Any]] = []
    for i in range(header_row + 1, ws.max_row + 1):
        rec: Dict[str, Any] = {}
        empty = True
        for j, key in headers_by_idx.items():
            v = ws.cell(row=i, column=j).value
            rec[key] = v
            if v not in (None, ""):
                empty = False
        if not empty:
            rows.append(rec)
    return rows


def _norm_keys(rec: Dict[str, Any]) -> Dict[str, Any]:
    # Normalize keys for fuzzy matching: lowercase and remove non-alphanumeric
    out = {}
    import re
    for k, v in rec.items():
        nk = re.sub(r"[^a-z0-9]", "", (str(k) or "").lower())
        out[nk] = v
    return out


def _get_first(nrec: Dict[str, Any], candidates: List[str]):
    for c in candidates:
        if c in nrec and nrec[c] not in (None, ""):
            return nrec[c]
    return None


def build_payload(rec: Dict[str, Any]) -> Dict[str, Any]:
    nrec = _norm_keys(rec)
    # Canonical OCR keys expected by API for board_certification
    ocr = {
        "abmsuid": _s(_get_first(nrec, [
            "boardabmsuid", "abmsuid"
        ])),
        "abms_name": _s(_get_first(nrec, [
            "providername", "provider_name", "abmsname", "boardabmsname"
        ])),
        "abms_dob": _s(_get_first(nrec, [
            "boardabmsdob", "dob"
        ])),
        "abms_education": _s(_get_first(nrec, [
            "boardabmseducation"
        ])),
        "abms_address": _s(_get_first(nrec, [
            "boardabmsaddress"
        ])),
        "abms_certification_board": _s(_get_first(nrec, [
            "boardabmscertificationboard"
        ])),
        "abms_certification_type": _s(_get_first(nrec, [
            "boardabmscertificationtype"
        ])),
        "abms_status": _s(_get_first(nrec, [
            "boardabmsstatus"
        ])),
        "abms_duration": _s(_get_first(nrec, [
            "boardabmsduration"
        ])),
        "abms_occurrence": _s(_get_first(nrec, [
            "boardabmsoccurrence"
        ])),
        "abms_start_date": _s(_get_first(nrec, [
            "boardabmsstartdate", "boardabmsstart_date"
        ])),
        "abms_end_date": _s(_get_first(nrec, [
            "boardabmsenddate", "boardabmsend_date"
        ])),
        "abms_reverification_date": _s(_get_first(nrec, [
            "boardabmsreverificationdate"
        ])),
        "abms_participating_in_moc": _s(_get_first(nrec, [
            "boardabmsparticipatinginmoc", "boardabmsparticipatinginmoc"
        ])),
    }
    # Drop None values
    ocr = {k: v for k, v in ocr.items() if v not in (None, "")}

    # Verification fields
    pdf_match = _s(_get_first(nrec, [
        "demo_verification_attribute1pdfformatmatch".replace("_", ""),
        "pdfformatmatch"
    ]))
    cert_match = _s(_get_first(nrec, [
        "demo_verification_attribute2boardcertificatematch".replace("_", ""),
        "boardcertificatematch"
    ]))
    cert_status = _s(_get_first(nrec, [
        "demo_verification_attribute3certificationstatus".replace("_", ""),
        "certificationstatus"
    ]))

    # Comments typically appear as Comment 1, Comment 2, Comment 3
    comment1 = _s(_get_first(nrec, ["comment1"]))
    comment2 = _s(_get_first(nrec, ["comment2"]))
    comment3 = _s(_get_first(nrec, ["comment3"]))

    verification = {
        "pdf_format_match": pdf_match,
        "board_certificate_match": cert_match,
        "certification_status": cert_status,
        "comment_1": comment1,
        "comment_2": comment2,
        "comment_3": comment3,
    }
    # Remove empty entries
    verification = {k: v for k, v in verification.items() if v not in (None, "")}

    # Extract npi for traceability
    npi_val = _get_first(nrec, ["npi", "providernpi", "provider_npi"])
    npi_str = None
    if npi_val is not None:
        try:
            npi_str = str(int(npi_val)) if isinstance(npi_val, (int, float)) else str(npi_val)
            npi_str = npi_str.strip()
        except Exception:
            npi_str = str(npi_val).strip()

    return {
        "source": f"{os.path.basename(EXCEL_PATH)}:{SHEET_NAME}",
        "npi": npi_str,
        "ocr": ocr,
        "verification": verification,
    }


def upsert_board_certification(db: Session, form_id: str, payload: Dict[str, Any]):
    row = (
        db.query(UploadedDocument)
        .filter(UploadedDocument.form_id == form_id)
        .filter(UploadedDocument.file_type == "board_certification")
        .first()
    )
    if not row:
        row = UploadedDocument(
            form_id=form_id,
            filename=f"board_cert_{form_id}.json",
            file_extension="json",
            file_type="board_certification",
            status="In Progress",
        )
        db.add(row)
        db.flush()
    row.ocr_output = json.dumps(payload.get("ocr", {}))
    row.verification_data = json.dumps(payload.get("verification", {}))
    if not row.status:
        row.status = "In Progress"
    return row


def main():
    rows = load_rows(EXCEL_PATH)
    print(f"Loaded Board_Certification rows: {len(rows)}")

    db: Session = SessionLocal()
    try:
        # Build NPI -> form_id index
        npi_to_form: Dict[str, str] = {}
        for f in db.query(FormData).all():
            if f.npi and f.form_id:
                npi_to_form[str(f.npi).strip()] = f.form_id
        for a in db.query(Application).all():
            if a.npi and a.form_id:
                npi_to_form.setdefault(str(a.npi).strip(), a.form_id)

        upserts = 0
        for rec in rows:
            # Determine form by NPI
            nrec = _norm_keys(rec)
            npi_val = None
            for cand in ("npi", "providernpi", "provider_npi"):
                if cand in nrec and nrec[cand] not in (None, ""):
                    npi_val = nrec[cand]
                    break
            if not npi_val:
                continue
            try:
                npi = str(int(npi_val)) if isinstance(npi_val, (int, float)) else str(npi_val)
                npi = npi.strip()
            except Exception:
                npi = str(npi_val).strip()
            form_id = npi_to_form.get(npi)
            if not form_id:
                continue
            payload = build_payload(rec)
            upsert_board_certification(db, form_id, payload)
            upserts += 1
        db.commit()
        print(f"Board Certification ingest complete. Upserts={upserts}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
