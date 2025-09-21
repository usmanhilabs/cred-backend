import os, sys, json
from typing import Any, Dict

# Ensure project root path
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

import openpyxl
from sqlalchemy.orm import Session
from app.database import SessionLocal
from app.models import UploadedDocument, Application, FormData

SHEET_NAME = "DEA"

OCR_HEADERS = [
    "Registrant Name",
    "DEA Registration Number",
    "Business Address",
    "Controlled Substance Schedules",
    "Business Activity",
    "Issue Date",
    "Expiration Date",
]
VERIF_HEADERS = [
    "Demo_Verification_Attribute1 (PDF Format Match)",
    "Comment 1",
    "Demo_Verification_Attribute2 (DEA Verification)",
]


def load_sheet(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in Excel file")
    ws = wb[SHEET_NAME]
    # Build header index map (lowercased)
    headers = {}
    header_row = None
    # find header row by scanning first 6 rows for a cell that equals 'Flag'
    for r in range(1, 7):
        row_vals = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[r]]
        if any(v == "flag" for v in row_vals):
            header_row = r
            break
    if not header_row:
        header_row = 2
    for j, cell in enumerate(ws[header_row], start=1):
        key = (str(cell.value) or "").strip()
        headers[j] = key

    # Parse rows after header_row
    rows = []
    for i in range(header_row + 1, ws.max_row + 1):
        rec = {}
        for j, key in headers.items():
            rec[key] = ws.cell(row=i, column=j).value
        raw_npi = rec.get("provider_npi") or rec.get("npi")
        if isinstance(raw_npi, (int, float)):
            npi = str(int(raw_npi))
        else:
            npi = str(raw_npi or "").strip()
        if not npi:
            continue
        rows.append(rec)
    return rows


def structured_payload(rec: Dict[str, Any]) -> Dict[str, Any]:
    # Build a structured dict with OCR and verification sections
    ocr = {}
    for h in OCR_HEADERS:
        ocr[h] = rec.get(h)
    ver = {
        "pdf_format_match": rec.get("Demo_Verification_Attribute1 (PDF Format Match)") or rec.get("PDF Format Match"),
        "comment_1": rec.get("Comment 1"),
        "dea_verification": rec.get("Demo_Verification_Attribute2 (DEA Verification)") or rec.get("DEA Verification"),
    }
    raw_npi = rec.get("provider_npi") or rec.get("npi")
    npi_str = str(int(raw_npi)) if isinstance(raw_npi, (int, float)) else str(raw_npi or "").strip()
    return {
        "source": "Other_Attributes_Schema.xlsx:DEA",
        "npi": npi_str,
        "ocr": ocr,
        "verification": ver,
    }


def upsert_dea(db: Session, form_id: str, payload: Dict[str, Any]):
    # Store as PSV type DEA in uploaded_documents
    row = (
        db.query(UploadedDocument)
        .filter(UploadedDocument.form_id == form_id)
        .filter(UploadedDocument.file_type == "DEA")
        .first()
    )
    if not row:
        row = UploadedDocument(
            form_id=form_id,
            filename=f"dea_{form_id}.json",
            file_extension="json",
            file_type="DEA",
            status="In Progress",
        )
        db.add(row)
        db.flush()
    row.ocr_output = json.dumps(payload.get("ocr", {}))
    row.verification_data = json.dumps(payload.get("verification", {}))
    if not row.status:
        row.status = "In Progress"
    return row


def main():
    excel_path = os.path.join(PROJECT_ROOT, "data", "Other_Attributes_Schema.xlsx")
    records = load_sheet(excel_path)
    print(f"Loaded DEA records: {len(records)}")

    db: Session = SessionLocal()
    try:
        npi_to_form = {}
        for f in db.query(FormData).all():
            if f.npi:
                npi_to_form[str(f.npi).strip()] = f.form_id
        for a in db.query(Application).all():
            if a.npi and a.form_id:
                npi_to_form.setdefault(str(a.npi).strip(), a.form_id)

        upserts = 0
        for rec in records:
            raw_npi = rec.get("provider_npi") or rec.get("npi")
            npi = str(int(raw_npi)) if isinstance(raw_npi, (int, float)) else str(raw_npi or "").strip()
            form_id = npi_to_form.get(npi)
            if not form_id:
                continue
            payload = structured_payload(rec)
            upsert_dea(db, form_id, payload)
            upserts += 1
        db.commit()
        print(f"DEA ingest complete. Upserts={upserts}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
