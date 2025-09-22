import os
import sys
import json
import re
from typing import Any, Dict, List, Tuple

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

import openpyxl  # type: ignore
from sqlalchemy.orm import Session
from app.database import SessionLocal
from app.models import UploadedDocument, Application, FormData

EXCEL_PATH = os.path.join(PROJECT_ROOT, "data", "BoardCertificate_License_Schema.xlsx")
SHEET_NAME = "License Board"


# Canonical OCR keys expected by the API for license_board
OCR_KEYS = [
    "LicenseBoard_ExtractedLicense",
    "LicenseBoard_Extracted_Name",
    "LicenseBoard_Extracted_License_Type",
    "LicenseBoard_Extracted_Primary_Status",
    "LicenseBoard_Extracted_Specialty",
    "LicenseBoard_Extracted_Qualification",
    "LicenseBoard_Extracted_School_Name",
    "LicenseBoard_Extracted_Graduation_Year",
    "LicenseBoard_Extracted_Previous_Names",
    "LicenseBoard_Extracted_Address",
    "LicenseBoard_Extracted_Issuance_Date",
    "LicenseBoard_Extracted_Expiration_Date",
    "LicenseBoard_Extracted_Current_Date_Time",
    "LicenseBoard_Extracted_Professional_Url",
    "LicenseBoard_Extracted_Disciplinary_Actions",
    "LicenseBoard_Extracted_Public_Record_Actions",
]

# Additional optional fields we may capture if present
EXTRA_OCR_KEYS = [
    "LicenseBoard_Extracted_Board_Name",
]


def _s(val: Any) -> Any:
    if val is None:
        return None
    try:
        import datetime as _dt
        if isinstance(val, (_dt.date, _dt.datetime)):
            return val.isoformat()
    except Exception:
        pass
    if isinstance(val, (str, int, float, bool)):
        return val
    return str(val)


def detect_header_row(ws) -> int:
    max_scan = min(8, ws.max_row)
    for r in range(1, max_scan + 1):
        vals = [(str(c.value) or "").strip().lower() for c in ws[r]]
        if any("licenseboard_extractedlicense" in v for v in vals) or any("npi" in v for v in vals):
            return r
    return 1


def load_rows(path: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in Excel file {path}")
    ws = wb[SHEET_NAME]
    header_row = detect_header_row(ws)
    # Capture headers and preserve duplicates with unique suffixes
    raw_headers_by_idx: Dict[int, str] = {}
    for j, cell in enumerate(ws[header_row], start=1):
        raw_headers_by_idx[j] = (str(cell.value) or "").strip()
    headers_by_idx: Dict[int, str] = {}
    seen_counts: Dict[str, int] = {}
    for j, name in raw_headers_by_idx.items():
        if name == "":
            name = f"__EMPTY_COL__{j}"
        count = seen_counts.get(name, 0)
        if count == 0:
            headers_by_idx[j] = name
        else:
            headers_by_idx[j] = f"{name}__{count+1}"
        seen_counts[name] = count + 1
    rows: List[Dict[str, Any]] = []
    for i in range(header_row + 1, ws.max_row + 1):
        rec: Dict[str, Any] = {}
        empty = True
        for j, key in headers_by_idx.items():
            v = ws.cell(row=i, column=j).value
            rec[key] = v
            if v not in (None, ""):
                empty = False
        if not empty:
            rows.append(rec)
    return rows


def build_payload(rec: Dict[str, Any]) -> Dict[str, Any]:
    # Normalize record keys for resilient lookups
    def norm_key(s: str) -> str:
        return re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")

    rec_norm: Dict[str, Any] = {norm_key(k): v for k, v in rec.items()}

    # Map Excel column variants (shown in the screenshot) to canonical OCR keys expected by API
    CANONICAL_OCR_MAP: Dict[str, List[str]] = {
        # License number and core fields
        "LicenseBoard_ExtractedLicense": [
            "license_website_license_number",
            "license_website_license_numbe",  # truncated header variant
            "licenseboard_extractedlicense",
        ],
        "LicenseBoard_Extracted_Name": [
            "license_website_name",  # often the provider name from the site
            "licenseboard_extracted_name",
        ],
        "LicenseBoard_Extracted_License_Type": [
            "license_website_licensetype",
            "license_board_licensetype",
            "licenseboard_extracted_license_type",
        ],
        "LicenseBoard_Extracted_Primary_Status": [
            "license_website_primarystatus",
            "licenseboard_extracted_primary_status",
        ],
        "LicenseBoard_Extracted_Specialty": [
            "license_website_special",
            "licenseboard_extracted_specialty",
        ],
        "LicenseBoard_Extracted_Qualification": [
            "license_website_qualification",
            "licenseboard_extracted_qualification",
        ],
        "LicenseBoard_Extracted_School_Name": [
            "license_website_schoolname",
            "licenseboard_extracted_school_name",
        ],
        "LicenseBoard_Extracted_Graduation_Year": [
            "license_website_graduationyear",
            "licenseboard_extracted_graduation_year",
        ],
        "LicenseBoard_Extracted_Previous_Names": [
            "license_website_previous_names",
            "licenseboard_extracted_previous_names",
        ],
        "LicenseBoard_Extracted_Address": [
            "license_website_address",
            "licenseboard_extracted_address",
        ],
        "LicenseBoard_Extracted_Issuance_Date": [
            "license_website_issuance_date",
            "licenseboard_extracted_issuance_date",
        ],
        "LicenseBoard_Extracted_Expiration_Date": [
            "license_website_expiration_date",
            "licenseboard_extracted_expiration_date",
        ],
        "LicenseBoard_Extracted_Current_Date_Time": [
            "license_website_currentdatetime",
            "licenseboard_extracted_current_date_time",
        ],
        "LicenseBoard_Extracted_Professional_Url": [
            "license_website_professionalurl",
            "licenseboard_extracted_professional_url",
        ],
        "LicenseBoard_Extracted_Disciplinary_Actions": [
            "license_website_disciplinaryactions",
            "licenseboard_extracted_disciplinary_actions",
        ],
        "LicenseBoard_Extracted_Public_Record_Actions": [
            "license_website_public_record_actions",
            "license_website_public_record_actions_",  # possible spacing variants
            "licenseboard_extracted_public_record_actions",
        ],
    }

    # Optional extras (won't break API)
    EXTRA_MAP: Dict[str, List[str]] = {
        "LicenseBoard_Extracted_Board_Name": [
            "license_website_boardname",
            "license_board_boardname",
        ]
    }

    def pick_first(keys: List[str]) -> Any:
        for k in keys:
            if k in rec_norm:
                v = rec_norm.get(k)
                if v not in (None, ""):
                    return v
        return None

    # Build OCR dict
    ocr: Dict[str, Any] = {}
    for canon, excel_keys in CANONICAL_OCR_MAP.items():
        ocr[canon] = _s(pick_first(excel_keys))
    for canon, excel_keys in EXTRA_MAP.items():
        val = pick_first(excel_keys)
        if val not in (None, ""):
            ocr[canon] = _s(val)

    # Verification label is handled in API; capture details present in the sheet
    def get_norm(*candidates: str) -> Any:
        for c in candidates:
            v = rec_norm.get(c)
            if v not in (None, ""):
                return v
        return None

    verification: Dict[str, Any] = {}
    verification["pdf_format_match"] = _s(
        get_norm(
            "demo_verification_attribute1_pdf_format_match",
            "pdf_format_match",
        )
    )
    verification["license_number_match"] = _s(
        get_norm(
            "demo_verification_attribute2_license_number_match",
            "license_number_match",
        )
    )
    verification["dates_match_status"] = _s(
        get_norm(
            "demo_verification_attribute3_dates_match_status",
            "dates_match_status",
        )
    )
    # Provider name match and score
    verification["provider_name_match"] = _s(get_norm("license_provider_match"))
    verification["provider_name_match_score"] = _s(get_norm("license_provider_match_score"))

    # Issuance/Expiration match flags: detect by relative position to corresponding date columns
    issuance_match = None
    expiration_match = None
    expect_issuance_next_match = False
    expect_expiration_next_match = False
    for k, v in rec.items():  # preserve original column order
        nk = norm_key(str(k))
        if nk in ("license_website_issuance_date",):
            expect_issuance_next_match = True
            continue
        if nk in ("license_website_expiration_date", "license_website_expirationdate"):
            expect_expiration_next_match = True
            continue
        if "match_with_input" in nk:
            if expect_issuance_next_match and issuance_match is None:
                issuance_match = v
                expect_issuance_next_match = False
                continue
            if expect_expiration_next_match and expiration_match is None:
                expiration_match = v
                expect_expiration_next_match = False
    verification["issuance_date_match"] = _s(issuance_match)
    verification["expiration_date_match"] = _s(expiration_match)

    # Misc note column sometimes present
    verification["expiry_issue"] = _s(
        get_norm("license_expiry_issue", "license_expiry_issue_")
    )

    # Comments: capture any 'comment' columns
    comments: List[Tuple[str, Any]] = []
    for k, v in rec.items():
        nk = norm_key(str(k))
        if nk.startswith("comment") and v not in (None, ""):
            comments.append((nk, v))
    # Sort deterministically and assign comment_1..comment_3 as available
    comments.sort(key=lambda x: x[0])
    for idx, (_, v) in enumerate(comments[:3], start=1):
        verification[f"comment_{idx}"] = _s(v)

    # Clean empties
    verification = {k: v for k, v in verification.items() if v not in (None, "")}

    # Trace NPI if available
    npi_raw = rec.get("npi") or rec.get("NPI") or rec.get("provider_npi")
    npi_str = None
    if npi_raw is not None:
        try:
            npi_str = str(int(npi_raw)) if isinstance(npi_raw, (int, float)) else str(npi_raw)
            npi_str = npi_str.strip()
        except Exception:
            npi_str = str(npi_raw).strip()

    return {
        "source": f"{os.path.basename(EXCEL_PATH)}:{SHEET_NAME}",
        "npi": npi_str,
        "ocr": ocr,
        "verification": verification,
    }


def upsert_license_board(db: Session, form_id: str, payload: Dict[str, Any]):
    row = (
        db.query(UploadedDocument)
        .filter(UploadedDocument.form_id == form_id)
        .filter(UploadedDocument.file_type == "license_board")
        .first()
    )
    if not row:
        row = UploadedDocument(
            form_id=form_id,
            filename=f"license_board_{form_id}.json",
            file_extension="json",
            file_type="license_board",
            status="In Progress",
        )
        db.add(row)
        db.flush()
    row.ocr_output = json.dumps(payload.get("ocr", {}))
    row.verification_data = json.dumps(payload.get("verification", {}))
    if not row.status:
        row.status = "In Progress"
    return row


def main():
    rows = load_rows(EXCEL_PATH)
    print(f"Loaded License Board rows: {len(rows)}")

    db: Session = SessionLocal()
    try:
        # Build NPI -> form_id index
        npi_to_form: Dict[str, str] = {}
        for f in db.query(FormData).all():
            if f.npi and f.form_id:
                npi_to_form[str(f.npi).strip()] = f.form_id
        for a in db.query(Application).all():
            if a.npi and a.form_id:
                npi_to_form.setdefault(str(a.npi).strip(), a.form_id)

        upserts = 0
        for rec in rows:
            # Prefer NPI linkage
            npi_raw = rec.get("npi") or rec.get("NPI") or rec.get("provider_npi")
            if not npi_raw:
                continue
            try:
                npi = str(int(npi_raw)) if isinstance(npi_raw, (int, float)) else str(npi_raw)
                npi = npi.strip()
            except Exception:
                npi = str(npi_raw).strip()
            form_id = npi_to_form.get(npi)
            if not form_id:
                continue
            payload = build_payload(rec)
            upsert_license_board(db, form_id, payload)
            upserts += 1
        db.commit()
        print(f"License Board ingest complete. Upserts={upserts}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
