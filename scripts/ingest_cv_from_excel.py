import os, sys, json
from typing import Any, Dict

# Ensure project root path
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

import openpyxl
from sqlalchemy.orm import Session
from app.database import SessionLocal
from app.models import UploadedDocument, Application, FormData

SHEET_NAME = "CV"

OCR_HEADERS = [
    "Provider Name",
    "Medical Education",
    "Postgraduate Training",
    "Board Certification",
    "Most Recent Work History",
]
VERIF_HEADERS = [
    "PDF Format Match",
    "Verification:",
]


def _npi_str(val: Any) -> str:
    if isinstance(val, (int, float)):
        return str(int(val))
    return str(val or "").strip()


def load_sheet(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in Excel file")
    ws = wb[SHEET_NAME]

    # Find header row by scanning first 6 rows for a cell equal to 'Flag' or 'Provider Name'
    headers = {}
    header_row = None
    for r in range(1, 8):
        row_vals = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        if any(v.lower() == "flag" for v in row_vals) or any(v == "Provider Name" for v in row_vals):
            header_row = r
            break
    if not header_row:
        header_row = 2

    # Map column index to exact header text
    for j, cell in enumerate(ws[header_row], start=1):
        key = (str(cell.value) or "").strip()
        headers[j] = key

    rows = []
    for i in range(header_row + 1, ws.max_row + 1):
        rec = {}
        for j, key in headers.items():
            rec[key] = ws.cell(row=i, column=j).value
        npi = _npi_str(rec.get("npi") or rec.get("provider_npi"))
        if not npi:
            continue
        rows.append(rec)
    return rows


def structured_payload(rec: Dict[str, Any]) -> Dict[str, Any]:
    ocr = {h: rec.get(h) for h in OCR_HEADERS}
    ver = {
        "pdf_format_match": rec.get("PDF Format Match"),
        "verification_summary": rec.get("Verification:") or rec.get("Verification")
    }
    return {
        "npi": _npi_str(rec.get("npi") or rec.get("provider_npi")),
        "ocr": ocr,
        "verification": ver,
        "source": "Other_Attributes_Schema.xlsx:CV",
    }


def upsert_cv(db: Session, form_id: str, payload: Dict[str, Any]):
    row = (
        db.query(UploadedDocument)
        .filter(UploadedDocument.form_id == form_id)
        .filter(UploadedDocument.file_type.in_(["CV", "cv", "cv/resume"]))
        .order_by(UploadedDocument.id.desc())
        .first()
    )
    if not row:
        row = UploadedDocument(
            form_id=form_id,
            filename=f"cv_{form_id}.json",
            file_extension="json",
            file_type="CV",
            status="In Progress",
        )
        db.add(row)
        db.flush()
    row.ocr_output = json.dumps(payload.get("ocr", {}))
    row.verification_data = json.dumps(payload.get("verification", {}))
    if not row.status:
        row.status = "In Progress"
    return row


def main():
    excel_path = os.path.join(PROJECT_ROOT, "data", "Other_Attributes_Schema.xlsx")
    records = load_sheet(excel_path)
    print(f"Loaded CV records: {len(records)}")

    db: Session = SessionLocal()
    try:
        npi_to_form = {}
        for f in db.query(FormData).all():
            if f.npi:
                npi_to_form[str(f.npi).strip()] = f.form_id
        for a in db.query(Application).all():
            if a.npi and a.form_id:
                npi_to_form.setdefault(str(a.npi).strip(), a.form_id)

        upserts = 0
        for rec in records:
            npi = _npi_str(rec.get("npi") or rec.get("provider_npi"))
            form_id = npi_to_form.get(npi)
            if not form_id:
                continue
            payload = structured_payload(rec)
            upsert_cv(db, form_id, payload)
            upserts += 1
        db.commit()
        print(f"CV ingest complete. Upserts={upserts}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
