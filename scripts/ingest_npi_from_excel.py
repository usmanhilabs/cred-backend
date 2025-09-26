import os
import sys
import json
from typing import Optional, Dict, Any

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.abspath(os.path.join(BASE_DIR, os.pardir))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from app.database import SessionLocal  # type: ignore
from app.models import Application, UploadedDocument, FormData  # type: ignore


def normalize_npi(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Handle scientific notation like 1.23456789E9
    try:
        if "e" in s.lower():
            num = float(s)
            s = str(int(num))
    except Exception:
        pass
    # Remove non-digits and pad to 10 if needed
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits:
        return None
    if len(digits) == 9:
        # Some sources miss a leading zero
        digits = digits.zfill(10)
    return digits

def upsert_npi_doc(session, form_id: str, verification_details: Dict[str, Any]):
    row = (
        session.query(UploadedDocument)
        .filter(UploadedDocument.form_id == form_id)
        .filter(UploadedDocument.file_type == "npi")
        .first()
    )
    if not row:
        row = UploadedDocument(
            form_id=form_id,
            filename=f"{verification_details['npi']}_NPI.png",
            file_extension="png",
            file_type="npi",
            status="In Progress",
        )
        session.add(row)
        session.flush()

    else:
        row.filename = f"{verification_details['npi']}_NPI.png"
        row.file_extension = "png"

    row.ocr_output = json.dumps({})
    row.verification_data = json.dumps(verification_details or {})
    if not row.status:
        row.status = "In Progress"


def main():
    excel_path = os.path.join(ROOT_DIR, "data", "Other_Attributes_Schema.xlsx")
    sheet_name = "NPI"

    if not os.path.exists(excel_path):
        print(f"Excel file not found at {excel_path}")
        sys.exit(1)

    wb = load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found in workbook")
        sys.exit(1)
    ws = wb[sheet_name]

    # Detect header row by scanning first 6 rows for any cell containing 'npi'
    header_row = None
    for r in range(1, min(7, ws.max_row + 1)):
        row_vals = [(str(c.value) or "").strip().lower() for c in ws[r]]
        if any("npi" in v for v in row_vals):
            header_row = r
            break
    if not header_row:
        header_row = 1

    headers_by_name: Dict[str, int] = {}
    headers_by_idx: Dict[int, str] = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        key = (str(cell.value) or "").strip()
        if key:
            headers_by_name[key.strip().lower()] = idx
            headers_by_idx[idx] = key

    # Find NPI column among various possible header names
    npi_col_idx = None
    for k, idx in headers_by_name.items():
        if "npi" in k:
            npi_col_idx = idx
            break
    if not npi_col_idx:
        print("Could not locate an NPI column in NPI sheet")
        sys.exit(1)

    session = SessionLocal()
    upserts = 0
    scanned = 0
    try:
        # Build NPI -> form_id map
        npi_to_form: Dict[str, str] = {}
        for f in session.query(FormData).all():
            if f.npi and f.form_id:
                npi_to_form[str(f.npi).strip()] = f.form_id
        for a in session.query(Application).all():
            if a.npi and a.form_id:
                npi_to_form.setdefault(str(a.npi).strip(), a.form_id)

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            scanned += 1
            raw_npi = row[npi_col_idx - 1]
            npi = normalize_npi(raw_npi)
            if not npi:
                continue

            form_id = npi_to_form.get(npi)
            if not form_id:
                continue

            # Optional fields: 'active', 'comment', 'deactivated date'
            def get_cell_by_name(name_options):
                for name in name_options:
                    name_l = name.lower()
                    if name_l in headers_by_name:
                        return row[headers_by_name[name_l] - 1]
                return None

            active_val = get_cell_by_name(["active", "is_active", "status"])
            comment_val = get_cell_by_name(["comment", "comments", "note"])
            deactivated_val = get_cell_by_name(["deactivated date", "deactivation date", "inactive since"])

            verification_details = {
                "source": "NPPES",
                "npi": npi,
                "verified": True if str(active_val).strip().lower() in {"yes", "true", "1", "active"} else False if active_val is not None else None,
                "deactivatedDate": str(deactivated_val) if deactivated_val else None,
                "comment": str(comment_val).strip() if comment_val else None,
            }

            upsert_npi_doc(session, form_id, verification_details)
            upserts += 1

        session.commit()
        print(f"NPI ingest complete. Rows scanned={scanned}, upserts={upserts}")
    except Exception as e:
        session.rollback()
        print(f"Error during NPI ingest: {e}")
        raise
    finally:
        session.close()


if __name__ == "__main__":
    main()
